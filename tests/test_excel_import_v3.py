#!/usr/bin/env python3
"""
批量上传设备功能v3修复方案测试
测试validate_device_data和import_devices_from_excel函数
"""
import os
import sys
import io
import pandas as pd
from io import BytesIO

# 添加项目根目录到Python路径
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.services.excel_service import validate_device_data, import_devices_from_excel, generate_device_template


def test_validate_device_data():
    """测试数据验证函数"""
    print("\n=== 测试 validate_device_data 函数 ===")

    # 测试1: 正常数据
    print("\n测试1: 正常数据验证")
    df_normal = pd.DataFrame({
        'hostname': ['SW-Test-01', 'SW-Test-02'],
        'ip_address': ['192.168.1.1', '192.168.1.2'],
        'vendor': ['华为', 'Cisco'],
        'model': ['S5735', 'Catalyst 9300'],
        'status': ['活跃', 'active'],
        'login_port': [22, 23]
    })
    devices, errors = validate_device_data(df_normal)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 2, f"期望2个设备，实际{len(devices)}"
    assert len(errors) == 0, f"期望0个错误，实际{len(errors)}"
    assert devices[0]['vendor'] == 'Huawei', f"厂商映射失败: {devices[0]['vendor']}"
    assert devices[0]['status'] == 'active', f"状态映射失败: {devices[0]['status']}"
    print("  ✅ 正常数据验证通过")

    # 测试2: 缺少必填字段
    print("\n测试2: 缺少必填字段验证")
    df_missing = pd.DataFrame({
        'hostname': ['SW-Test-01', ''],
        'ip_address': ['192.168.1.1', '192.168.1.2'],
        'vendor': ['华为', '思科'],
        'model': ['S5735', 'Catalyst 9300']
    })
    devices, errors = validate_device_data(df_missing)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 1, f"期望1个设备，实际{len(devices)}"
    assert len(errors) == 1, f"期望1个错误，实际{len(errors)}"
    assert '缺少必填字段' in errors[0], f"错误信息不正确: {errors[0]}"
    print("  ✅ 必填字段验证通过")

    # 测试3: IP地址格式验证
    print("\n测试3: IP地址格式验证")
    df_ip = pd.DataFrame({
        'hostname': ['SW-Test-01', 'SW-Test-02'],
        'ip_address': ['192.168.1.1', 'invalid_ip'],
        'vendor': ['华为', '思科'],
        'model': ['S5735', 'Catalyst 9300']
    })
    devices, errors = validate_device_data(df_ip)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 1, f"期望1个设备，实际{len(devices)}"
    assert len(errors) == 1, f"期望1个错误，实际{len(errors)}"
    assert 'IP地址格式无效' in errors[0], f"错误信息不正确: {errors[0]}"
    print("  ✅ IP地址格式验证通过")

    # 测试4: 端口范围验证（v3优化）
    print("\n测试4: 端口范围验证（v3优化）")
    df_port = pd.DataFrame({
        'hostname': ['SW-Test-01', 'SW-Test-02'],
        'ip_address': ['192.168.1.1', '192.168.1.2'],
        'vendor': ['华为', '思科'],
        'model': ['S5735', 'Catalyst 9300'],
        'login_port': [0, 70000]  # 超出范围
    })
    devices, errors = validate_device_data(df_port)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 2, f"期望2个设备，实际{len(devices)}"
    assert len(errors) == 2, f"期望2个警告，实际{len(errors)}"
    assert devices[0]['login_port'] == 22, f"端口未修正: {devices[0]['login_port']}"
    assert '端口号 0 超出范围' in errors[0], f"错误信息不正确: {errors[0]}"
    assert '端口号 70000 超出范围' in errors[1], f"错误信息不正确: {errors[1]}"
    print("  ✅ 端口范围验证通过")

    # 测试5: Telnet端口自动修正
    print("\n测试5: Telnet端口自动修正")
    df_telnet = pd.DataFrame({
        'hostname': ['SW-Test-01'],
        'ip_address': ['192.168.1.1'],
        'vendor': ['华为'],
        'model': ['S5735'],
        'login_method': ['telnet'],
        'login_port': [22]
    })
    devices, errors = validate_device_data(df_telnet)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 1, f"期望1个设备，实际{len(devices)}"
    assert devices[0]['login_port'] == 23, f"Telnet端口未修正: {devices[0]['login_port']}"
    assert 'Telnet登录方式下端口自动修正为23' in errors[0], f"错误信息不正确: {errors[0]}"
    print("  ✅ Telnet端口自动修正通过")

    # 测试6: 中文列名映射
    print("\n测试6: 中文列名映射")
    df_cn = pd.DataFrame({
        '主机名': ['SW-Test-01'],
        'IP地址': ['192.168.1.1'],
        '厂商': ['华为'],
        '型号': ['S5735'],
        '状态': ['维护']
    })
    devices, errors = validate_device_data(df_cn)
    print(f"  有效设备数: {len(devices)}, 错误数: {len(errors)}")
    assert len(devices) == 1, f"期望1个设备，实际{len(devices)}"
    assert devices[0]['status'] == 'maintenance', f"状态映射失败: {devices[0]['status']}"
    print("  ✅ 中文列名映射通过")

    print("\n✅ 所有 validate_device_data 测试通过!")
    return True


def test_generate_template():
    """测试模板生成功能"""
    print("\n=== 测试 generate_device_template 函数 ===")

    # 测试1: 生成模板
    print("\n测试1: 生成模板")
    template = generate_device_template()
    assert template is not None, "模板生成失败"
    template_size = len(template.getvalue())
    assert template_size > 0, "模板大小为0"
    print(f"  模板生成成功，大小: {template_size} 字节")

    # 测试2: 验证模板内容
    print("\n测试2: 验证模板内容")
    template.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(template)

    # 验证工作表
    assert '设备清单' in wb.sheetnames, f"缺少'设备清单'工作表: {wb.sheetnames}"
    assert '填写说明' in wb.sheetnames, f"缺少'填写说明'工作表: {wb.sheetnames}"
    print(f"  工作表: {wb.sheetnames}")

    # 验证设备清单表头
    ws_data = wb['设备清单']
    headers = [cell.value for cell in ws_data[1]]
    print(f"  表头: {headers}")
    assert '主机名*' in headers, "缺少中文表头"
    assert 'IP地址*' in headers, "缺少IP地址表头"
    assert '厂商*' in headers, "缺少厂商表头"

    # 验证示例数据
    print(f"  数据行数: {ws_data.max_row - 1}")
    assert ws_data.max_row > 1, "缺少示例数据"

    # 验证填写说明工作表
    ws_help = wb['填写说明']
    help_headers = [cell.value for cell in ws_help[1]]
    print(f"  说明表头: {help_headers}")
    assert '安全提示' in help_headers, "v3优化: 缺少安全提示列"

    # 验证安全提示区域
    security_found = False
    for row in ws_help.iter_rows(values_only=True):
        if row and '安全提示' in str(row[0]):
            security_found = True
            break
    assert security_found, "v3优化: 缺少安全提示区域"
    print("  ✅ 安全提示区域验证通过")

    print("\n✅ 所有 generate_device_template 测试通过!")
    return True


def test_import_devices_mock():
    """测试导入设备功能（使用mock数据库会话）"""
    print("\n=== 测试 import_devices_from_excel 函数（Mock测试）===")

    # 创建一个简单的mock会话类
    class MockSession:
        def __init__(self):
            self.devices = []
            self.committed = False
            self.rolled_back = False

        def query(self, model):
            class MockQuery:
                def __init__(self, devices):
                    self.devices = devices

                def filter(self, *conditions):
                    return self

                def first(self):
                    return None  # 模拟设备不存在

                def all(self):
                    return []

                def filter_by(self, **kwargs):
                    return self
            return MockQuery(self.devices)

        def add(self, obj):
            self.devices.append(obj)

        def commit(self):
            self.committed = True

        def rollback(self):
            self.rolled_back = True

        def bulk_insert_mappings(self, model, mappings):
            for m in mappings:
                self.devices.append(m)

    # 测试数据
    df_test = pd.DataFrame({
        'hostname': ['SW-Test-01', 'SW-Test-02'],
        'ip_address': ['192.168.1.101', '192.168.1.102'],
        'vendor': ['华为', '思科'],
        'model': ['S5735', 'Catalyst 9300'],
        'status': ['活跃', '维护']
    })

    # 转换为Excel文件
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_test.to_excel(writer, index=False, sheet_name='设备清单')
    excel_buffer.seek(0)

    # 测试导入
    print("\n测试: 导入新设备")
    mock_session = MockSession()
    stats = import_devices_from_excel(excel_buffer, mock_session, skip_existing=False)

    print(f"  统计结果: {stats}")
    assert stats['total'] == 2, f"期望总数2，实际{stats['total']}"
    assert stats['success'] == 2, f"期望成功2，实际{stats['success']}"
    assert stats['failed'] == 0, f"期望失败0，实际{stats['failed']}"
    assert len(stats['errors']) == 0, f"期望0个错误，实际{len(stats['errors'])}"

    print("\n✅ import_devices_from_excel Mock测试通过!")
    return True


def run_all_tests():
    """运行所有测试"""
    print("=" * 60)
    print("批量上传设备功能v3修复方案测试")
    print("=" * 60)

    all_passed = True

    try:
        if not test_validate_device_data():
            all_passed = False
    except Exception as e:
        print(f"\n❌ validate_device_data 测试失败: {e}")
        import traceback
        traceback.print_exc()
        all_passed = False

    try:
        if not test_generate_template():
            all_passed = False
    except Exception as e:
        print(f"\n❌ generate_device_template 测试失败: {e}")
        import traceback
        traceback.print_exc()
        all_passed = False

    try:
        if not test_import_devices_mock():
            all_passed = False
    except Exception as e:
        print(f"\n❌ import_devices_from_excel 测试失败: {e}")
        import traceback
        traceback.print_exc()
        all_passed = False

    print("\n" + "=" * 60)
    if all_passed:
        print("✅ 所有测试通过!")
    else:
        print("❌ 部分测试失败!")
    print("=" * 60)

    return all_passed


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
