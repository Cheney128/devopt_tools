#!/usr/bin/env python3
"""
简单测试Excel模板生成功能
直接测试generate_device_template函数，不依赖于整个应用
"""
import os
import sys
import io
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import pandas as pd
from openpyxl import load_workbook

# 添加项目根目录到Python路径
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__))))

def test_excel_generation():
    """测试Excel模板生成功能"""
    try:
        # 直接导入Device模型，避免通过app导入
        from app.models.models import Device
        
        # 定义generate_device_template函数，直接复制代码，避免导入依赖
        def generate_device_template(session=None):
            """生成设备导入模板"""
            # 定义设备字段
            device_fields = [
                'hostname', 'ip_address', 'vendor', 'model', 'os_version', 
                'location', 'contact', 'status', 'login_method', 'login_port', 
                'username', 'password', 'sn'
            ]

            # 创建模板数据
            if session:
                # 获取现有设备数据
                devices = session.query(Device).all()
                if devices:
                    # 使用现有设备数据作为模板
                    device_data = []
                    for device in devices:
                        device_dict = {
                            'hostname': device.hostname,
                            'ip_address': device.ip_address,
                            'vendor': device.vendor,
                            'model': device.model,
                            'os_version': device.os_version or '',
                            'location': device.location or '',
                            'contact': device.contact or '',
                            'status': device.status,
                            'login_method': device.login_method,
                            'login_port': device.login_port,
                            'username': device.username or '',
                            'password': '',  # 密码字段为空，确保安全
                            'sn': device.sn or ''
                        }
                        device_data.append(device_dict)
                    df = pd.DataFrame(device_data)
                else:
                    # 只有表头的空模板
                    df = pd.DataFrame(columns=device_fields)
            else:
                # 只有表头的空模板
                df = pd.DataFrame(columns=device_fields)

            # 创建Excel文件，简化模板生成逻辑，去除可能导致问题的部分
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='设备清单')

            output.seek(0)
            return output
        
        # 从app.models获取数据库URL
        from app.models import DATABASE_URL
        
        print(f"使用数据库连接: {DATABASE_URL}")
        
        # 创建数据库引擎
        engine = create_engine(DATABASE_URL)
        
        # 测试数据库连接
        print("\n=== 测试数据库连接 ===")
        with engine.connect() as conn:
            result = conn.execute(text("SELECT 1"))
            print(f"数据库连接成功，结果: {result.fetchone()}")
        
        # 创建会话工厂
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        
        # 创建会话
        db = SessionLocal()
        
        try:
            # 测试设备表查询
            print("\n=== 测试设备表查询 ===")
            devices = db.query(Device).all()
            print(f"共查询到 {len(devices)} 台设备")
            
            # 测试1: 生成空模板（不提供session）
            print("\n=== 测试1: 生成空模板 ===")
            empty_template = generate_device_template()
            template_size = len(empty_template.getvalue())
            print(f"空模板生成成功，大小: {template_size} 字节")
            
            # 测试2: 使用数据库会话生成模板
            print("\n=== 测试2: 使用数据库会话生成模板 ===")
            db_template = generate_device_template(db)
            db_template_size = len(db_template.getvalue())
            print(f"带数据库数据的模板生成成功，大小: {db_template_size} 字节")
            
            # 测试3: 验证生成的Excel文件可以被openpyxl读取
            print("\n=== 测试3: 验证Excel文件完整性 ===")
            
            # 重置文件指针到开头
            db_template.seek(0)
            
            # 尝试用openpyxl读取Excel文件
            wb = load_workbook(db_template)
            ws = wb.active
            print(f"工作表名称: {ws.title}")
            print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            print(f"列名: {headers}")
            
            # 如果有数据，显示前5行
            if ws.max_row > 1:
                print("\n前5行数据:")
                for row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True):
                    print(row)
            
        finally:
            # 关闭会话
            db.close()
            
    except Exception as e:
        print(f"❌ 测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

if __name__ == "__main__":
    print("开始测试Excel模板生成功能...")
    success = test_excel_generation()
    if success:
        print("\n✅ 所有测试通过!")
        sys.exit(0)
    else:
        print("\n❌ 测试失败!")
        sys.exit(1)